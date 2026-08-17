"""Populate workbook poster cells from existing TMDB IDs or directories."""

from __future__ import annotations

import argparse
import os
import re
import shutil
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_FILES = [
    "MCU.xlsx",
    "STARWARS.xlsx",
    "STARTREK.xlsx",
    "DOCTORWHO.xlsx",
    "MIDDLEEARTH.xlsx",
    "RIORDANVERSE.xlsx",
    "DCU.xlsx",
    "ZELDA.xlsx",
    "ANDYWEIR.xlsx",
    "BIGBANGTHEORY.xlsx",
]
API_BASE = "https://api.themoviedb.org/3"
IMAGE_BASE = "https://image.tmdb.org/t/p/w500"


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def text(value: object) -> str:
    return str(value or "").strip()


def get_value(row_map: dict[str, object], *keys: str) -> str:
    for key in keys:
        value = text(row_map.get(normalize(key), ""))
        if value:
            return value
    return ""


def parse_reference(value: str, default_type: str) -> tuple[str, str, str | None]:
    raw = value.strip()
    if not raw:
        return "", "", None

    path = urlparse(raw).path if raw.startswith("http") else raw
    season_match = re.search(r"/(tv)/(\d+)/season/(\d+)", path, re.I)
    if season_match:
        return season_match.group(1).lower(), season_match.group(2), season_match.group(3)

    media_match = re.search(r"/(tv|movie)/(\d+)", path, re.I)
    if media_match:
        return media_match.group(1).lower(), media_match.group(2), None

    numeric_match = re.search(r"\d+", raw)
    if numeric_match:
        return default_type, numeric_match.group(0), None
    return "", "", None


def infer_type(type_value: str, explicit_type: str) -> str:
    explicit = explicit_type.lower().strip()
    if explicit in {"tv", "movie"}:
        return explicit
    key = type_value.lower()
    return "tv" if any(token in key for token in ("tv", "series", "episode", "special", "short")) else "movie"


def fetch_json(session: requests.Session, token: str, path: str) -> dict:
    response = session.get(
        f"{API_BASE}{path}",
        params={"language": "en-US"},
        headers={"Authorization": f"Bearer {token}", "Accept": "application/json"},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def process_workbook(path: Path, session: requests.Session, token: str, write_posters: bool, overwrite: bool) -> tuple[int, int, int]:
    workbook = load_workbook(path, data_only=False)
    changed = False
    processed = written = skipped = 0

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=False))
        if not rows:
            continue
        headers = [text(cell.value) for cell in rows[0]]
        keys = [normalize(value) for value in headers]
        poster_index = next((i for i, key in enumerate(keys) if key == "poster"), None)
        if poster_index is None:
            continue

        for row_number, cells in enumerate(rows[1:], start=2):
            row_map = {keys[i]: cells[i].value for i in range(min(len(keys), len(cells)))}
            existing_poster = get_value(row_map, "poster")
            if existing_poster and not overwrite:
                skipped += 1
                continue

            type_value = get_value(row_map, "type")
            explicit_type = get_value(row_map, "tmdb type", "tmdbtype", "tmdb_type")
            reference = get_value(
                row_map,
                "tmdb directory",
                "tmdb path",
                "tmdb url",
                "tmdb id",
                "tmdbid",
                "tmdb_id",
                "themoviedb directory",
                "themoviedb path",
                "themoviedb url",
                "themoviedb id",
            )
            media_type = infer_type(type_value, explicit_type)
            media_type, tmdb_id, season_number = parse_reference(reference, media_type)
            if not tmdb_id:
                continue

            processed += 1
            try:
                details = fetch_json(session, token, f"/{media_type}/{tmdb_id}")
                poster_path = ""
                if media_type == "tv" and season_number is not None:
                    season = fetch_json(session, token, f"/tv/{tmdb_id}/season/{season_number}")
                    poster_path = text(season.get("poster_path"))
                if not poster_path:
                    poster_path = text(details.get("poster_path"))
                if not poster_path:
                    continue

                cells[poster_index].value = f"{IMAGE_BASE}/{poster_path.lstrip('/')}"
                changed = True
                written += 1
            except requests.RequestException as exc:
                print(f"TMDB request failed: {path.name} / {worksheet.title} row {row_number}: {exc}")

    if changed:
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        workbook.save(path)
    return processed, written, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--token", default=os.getenv("TMDB_API_TOKEN"), help="TMDB API Read Access Token; defaults to TMDB_API_TOKEN")
    parser.add_argument("--write-posters", action="store_true", help="Write remote TMDB image URLs into poster cells")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing poster values")
    args = parser.parse_args()

    if not args.token:
        raise SystemExit("Provide --token or set TMDB_API_TOKEN. The token is never stored in index.html.")

    session = requests.Session()
    totals = [0, 0, 0]
    for filename in WORKBOOK_FILES:
        path = ROOT / filename
        if not path.exists():
            continue
        values = process_workbook(path, session, args.token, args.write_posters, args.overwrite)
        totals = [left + right for left, right in zip(totals, values)]

    print(f"Rows with TMDB references processed: {totals[0]}")
    print(f"Remote poster URLs {'written' if args.write_posters else 'available'}: {totals[1]}")
    print(f"Rows skipped because they already had posters: {totals[2]}")
    if not args.write_posters:
        print("Dry run only. Add --write-posters after reviewing the affected rows.")


if __name__ == "__main__":
    main()
