from openpyxl import load_workbook

path = r"c:\Users\benja\timeline-site\STARWARS.xlsx"
wb = load_workbook(path)
ws = wb["CANON"]
headers = [c.value for c in ws[1]]
poster_idx = headers.index("poster")
cv_idx = headers.index("comic vine id")

rows_with_cv = []
rows_with_poster_but_no_cv = []
rows_with_http_poster = []
nonblank_poster_count = 0
for row in ws.iter_rows(min_row=2, values_only=False):
    cv = row[cv_idx].value if cv_idx < len(row) else None
    poster = row[poster_idx].value if poster_idx < len(row) else None

    if cv is not None and str(cv).strip():
        rows_with_cv.append((row[3].value, poster))
    if poster is not None and str(poster).strip() and (cv is None or not str(cv).strip()):
        rows_with_poster_but_no_cv.append((row[3].value, poster))
    if poster is not None and str(poster).strip() and str(poster).lower().startswith("http"):
        rows_with_http_poster.append((row[3].value, poster))
        nonblank_poster_count += 1

print("rows_with_cv_value", len(rows_with_cv))
print("rows_with_http_poster", len(rows_with_http_poster))
print("rows_with_poster_but_no_cv", len(rows_with_poster_but_no_cv))
print("sample_cv_rows", rows_with_cv[:10])
print("sample_poster_without_cv", rows_with_poster_but_no_cv[:10])
