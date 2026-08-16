import csv
import json
import re
import time
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_CSV = ROOT / "tmdb_id_suggestions.csv"
OUTPUT_JSON = ROOT / "tmdb_id_suggestions.json"

WORKBOOK_FILES = [
    "MCU.xlsx",
    "STARWARS.xlsx",
    "STARTREK.xlsx",
    "DOCTORWHO.xlsx",
    "ANDYWEIR.xlsx",
    "MIDDLEEARTH.xlsx",
    "RIORDANVERSE.xlsx",
    "DCU.xlsx",
    "BIGBANGTHEORY.xlsx",
    "ZELDA.xlsx",
]

TMDB_API_BASE = "https://api.themoviedb.org/3"


def normalize(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def get_value(row_map, *keys):
    for key in keys:
        v = row_map.get(normalize(key), "")
        if v is None:
            continue
        text = str(v).strip()
        if text:
            return text
    return ""


def infer_media_type(type_text):
    key = str(type_text or "").strip().lower()
    if any(token in key for token in ["tv", "series", "episode", "special", "doctor who"]):
        return "tv"
    return "movie"


def parse_year(value):
    text = str(value or "").strip()
    if not text:
        return None

    m = re.search(r"\b(19|20)\d{2}\b", text)
    if m:
        return int(m.group(0))

    try:
        dt = datetime.fromisoformat(text.replace("Z", ""))
        return dt.year
    except Exception:
        return None


def parse_token_from_index_html(index_html_path):
    content = index_html_path.read_text(encoding="utf-8", errors="ignore")
    m = re.search(r"bearerToken:\s*'([^']+)'", content)
    if not m:
        return ""
    return m.group(1).strip()


def get_result_title(item, media_type):
    if media_type == "tv":
        return (item.get("name") or item.get("original_name") or "").strip()
    return (item.get("title") or item.get("original_title") or "").strip()


def get_result_date(item, media_type):
    if media_type == "tv":
        return (item.get("first_air_date") or "").strip()
    return (item.get("release_date") or "").strip()


def score_candidate(query_title, query_year, item_title, item_year, popularity):
    if not item_title:
        return 0.0

    text_score = SequenceMatcher(None, query_title.lower(), item_title.lower()).ratio()

    year_score = 0.0
    if query_year and item_year:
        diff = abs(query_year - item_year)
        year_score = max(0.0, 1.0 - (diff / 6.0))
    elif not query_year:
        year_score = 0.5

    pop_score = min(float(popularity or 0.0), 200.0) / 200.0

    return (0.72 * text_score) + (0.23 * year_score) + (0.05 * pop_score)


def search_tmdb(session, token, media_type, title, year):
    url = f"{TMDB_API_BASE}/search/{media_type}"
    params = {
        "query": title,
        "include_adult": "false",
        "language": "en-US",
    }
    if year:
        if media_type == "tv":
            params["first_air_date_year"] = str(year)
        else:
            params["year"] = str(year)

    response = session.get(
        url,
        params=params,
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return payload.get("results") or []


def process_workbook(path, session, token, cache):
    suggestions = []
    wb = load_workbook(path, data_only=True, read_only=True)

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        header_keys = [normalize(h) for h in header]

        for row_number, row in enumerate(rows, start=2):
            row_map = {header_keys[i]: row[i] for i in range(min(len(header_keys), len(row)))}

            existing_id = get_value(row_map, "tmdb id", "tmdbid", "tmdb_id", "themoviedb id")
            existing_type = get_value(row_map, "tmdb type", "tmdbtype", "tmdb_type")
            if existing_id:
                suggestions.append(
                    {
                        "workbook": path.name,
                        "sheet": ws.title,
                        "row_number": row_number,
                        "title": get_value(row_map, "episode title", "title", "serial title"),
                        "type": get_value(row_map, "type"),
                        "release_date": get_value(row_map, "release date"),
                        "query_title": "",
                        "media_type": existing_type or infer_media_type(get_value(row_map, "type")),
                        "tmdb_id": existing_id,
                        "tmdb_type": existing_type or infer_media_type(get_value(row_map, "type")),
                        "matched_title": "(existing)",
                        "matched_date": "",
                        "score": 1.0,
                        "status": "existing",
                    }
                )
                continue

            query_title = get_value(row_map, "episode title", "title", "serial title")
            if not query_title:
                continue

            type_text = get_value(row_map, "type")
            media_type = infer_media_type(type_text)
            release_date = get_value(row_map, "release date")
            query_year = parse_year(release_date)

            cache_key = (media_type, query_title.lower(), query_year)
            if cache_key not in cache:
                try:
                    cache[cache_key] = search_tmdb(session, token, media_type, query_title, query_year)
                except Exception:
                    cache[cache_key] = []
                time.sleep(0.06)

            results = cache[cache_key]
            best = None
            best_score = 0.0

            for item in results[:8]:
                item_title = get_result_title(item, media_type)
                item_date = get_result_date(item, media_type)
                item_year = parse_year(item_date)
                score = score_candidate(
                    query_title=query_title,
                    query_year=query_year,
                    item_title=item_title,
                    item_year=item_year,
                    popularity=item.get("popularity"),
                )
                if score > best_score:
                    best_score = score
                    best = (item, item_title, item_date)

            if best is None:
                status = "no_match"
                tmdb_id = ""
                matched_title = ""
                matched_date = ""
                tmdb_type = media_type
                final_score = 0.0
            else:
                status = "matched" if best_score >= 0.62 else "review"
                tmdb_id = str(best[0].get("id") or "")
                tmdb_type = media_type
                matched_title = best[1]
                matched_date = best[2]
                final_score = round(best_score, 4)

            suggestions.append(
                {
                    "workbook": path.name,
                    "sheet": ws.title,
                    "row_number": row_number,
                    "title": query_title,
                    "type": type_text,
                    "release_date": release_date,
                    "query_title": query_title,
                    "media_type": media_type,
                    "tmdb_id": tmdb_id,
                    "tmdb_type": tmdb_type,
                    "matched_title": matched_title,
                    "matched_date": matched_date,
                    "score": final_score,
                    "status": status,
                }
            )

    return suggestions


def main():
    token = parse_token_from_index_html(ROOT / "index.html")
    if not token:
        raise SystemExit("No TMDB bearer token found in index.html tmdbConfig.bearerToken")

    session = requests.Session()
    cache = {}
    all_rows = []

    for file_name in WORKBOOK_FILES:
        workbook_path = ROOT / file_name
        if not workbook_path.exists():
            continue
        all_rows.extend(process_workbook(workbook_path, session, token, cache))

    OUTPUT_CSV.write_text("", encoding="utf-8")
    fieldnames = [
        "workbook",
        "sheet",
        "row_number",
        "title",
        "type",
        "release_date",
        "query_title",
        "media_type",
        "tmdb_id",
        "tmdb_type",
        "matched_title",
        "matched_date",
        "score",
        "status",
    ]

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    with OUTPUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)

    total = len(all_rows)
    matched = sum(1 for row in all_rows if row["status"] in ("matched", "existing"))
    review = sum(1 for row in all_rows if row["status"] == "review")
    no_match = sum(1 for row in all_rows if row["status"] == "no_match")

    print(f"Generated {total} TMDB ID suggestions")
    print(f"Matched/existing: {matched}")
    print(f"Needs review: {review}")
    print(f"No match: {no_match}")
    print(f"CSV: {OUTPUT_CSV}")
    print(f"JSON: {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
