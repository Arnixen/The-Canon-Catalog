"""Populate poster cells from exact Comic Vine issue IDs."""

from __future__ import annotations

import argparse
import os
import re
import shutil
import time
from collections import deque
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "STARWARS.xlsx"
API_BASE = "https://comicvine.gamespot.com/api"
USER_AGENT = "The-Canon-Catalog Comic Vine ID poster resolver/1.0"


def text(value: object) -> str:
    return str(value or "").strip()


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def parse_issue_id(value: object) -> str:
    raw = text(value)
    if raw.isdigit():
        return raw
    match = re.search(r"(?:^|[-/])(\d+)(?:/)?(?:\?.*)?$", raw)
    return match.group(1) if match else ""


def fetch_issue(session: requests.Session, api_key: str, issue_id: str) -> dict:
    response = session.get(
        f"{API_BASE}/issue/4000-{issue_id}/",
        params={"api_key": api_key, "format": "json"},
        headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
        timeout=30,
    )
    if response.status_code == 429:
        raise RuntimeError("Comic Vine rate limit reached (HTTP 429)")
    response.raise_for_status()
    payload = response.json()
    if payload.get("status_code") not in (None, 1):
        raise RuntimeError(payload.get("error") or "Comic Vine request failed")
    return payload.get("results") or {}


def process_workbook(
    path: Path,
    api_key: str,
    write_posters: bool,
    overwrite: bool,
    delay: float,
    max_requests_per_hour: int,
) -> tuple[int, int, int, int]:
    workbook = load_workbook(path, data_only=False)
    session = requests.Session()
    cache: dict[str, str] = {}
    processed = written = skipped = failed = 0
    changed = False
    request_times: deque[float] = deque()

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=False))
        if not rows:
            continue
        headers = [normalize(cell.value) for cell in rows[0]]
        id_index = next((i for i, key in enumerate(headers) if key == "comicvineid"), None)
        poster_index = next((i for i, key in enumerate(headers) if key == "poster"), None)
        if id_index is None or poster_index is None:
            continue

        for row_number, cells in enumerate(rows[1:], start=2):
            issue_id = parse_issue_id(cells[id_index].value)
            if not issue_id:
                continue
            if text(cells[poster_index].value) and not overwrite:
                skipped += 1
                continue

            processed += 1
            if issue_id not in cache:
                try:
                    now = time.monotonic()
                    while request_times and now - request_times[0] >= 3600:
                        request_times.popleft()
                    if len(request_times) >= max_requests_per_hour:
                        wait = 3600 - (now - request_times[0])
                        print(f"RATE_LIMIT waiting {wait:.1f}s before the next API request")
                        time.sleep(wait)
                        request_times.popleft()
                    if request_times:
                        elapsed = time.monotonic() - request_times[-1]
                        if elapsed < delay:
                            time.sleep(delay - elapsed)
                    request_times.append(time.monotonic())
                    item = fetch_issue(session, api_key, issue_id)
                    cache[issue_id] = text((item.get("image") or {}).get("original_url"))
                except (requests.RequestException, RuntimeError) as exc:
                    failed += 1
                    print(f"FAILED {worksheet.title}!{row_number} issue {issue_id}: {exc}")
                    continue

            image_url = cache[issue_id]
            if not image_url:
                failed += 1
                print(f"NO_COVER {worksheet.title}!{row_number} issue {issue_id}")
                continue

            print(f"FOUND {worksheet.title}!{row_number} issue {issue_id}: {image_url}")
            if write_posters:
                cells[poster_index].value = image_url
                written += 1
                changed = True
                print(f"WROTE {worksheet.title}!{row_number} issue {issue_id}: {image_url}")

    if changed:
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        workbook.save(path)
    return processed, written, skipped, failed


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--api-key", default=os.getenv("COMIC_VINE_API_KEY"))
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--write-posters", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--delay", type=float, default=1.1, help="Minimum seconds between unique API requests")
    parser.add_argument("--max-requests-per-hour", type=int, default=199, help="Hourly request cap below Comic Vine's 200-request limit")
    args = parser.parse_args()

    if not args.api_key:
        raise SystemExit("Provide --api-key or set COMIC_VINE_API_KEY")

    processed, written, skipped, failed = process_workbook(
        args.workbook,
        args.api_key,
        args.write_posters,
        args.overwrite,
        args.delay,
        args.max_requests_per_hour,
    )
    print(f"Rows with Comic Vine IDs processed: {processed}")
    print(f"Poster URLs {'written' if args.write_posters else 'found'}: {written if args.write_posters else processed - failed}")
    print(f"Rows skipped because they already had posters: {skipped}")
    print(f"Lookup failures: {failed}")
    if not args.write_posters:
        print("Dry run only. Add --write-posters after reviewing the results.")


if __name__ == "__main__":
    main()