"""Suggest and optionally write published-book covers from Open Library."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import time
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_FILES = [
    "MCU.xlsx", "STARWARS.xlsx", "STARTREK.xlsx", "DOCTORWHO.xlsx",
    "ANDYWEIR.xlsx", "MIDDLEEARTH.xlsx", "RIORDANVERSE.xlsx", "DCU.xlsx",
    "BIGBANGTHEORY.xlsx", "ZELDA.xlsx",
]
SEARCH_URL = "https://openlibrary.org/search.json"
COVER_URL = "https://covers.openlibrary.org/b/id/{cover_id}-L.jpg"
USER_AGENT = "The-Canon-Catalog Open Library cover resolver/1.0"
YEAR_TOLERANCE = 5


def text(value: object) -> str:
    return str(value or "").strip()


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def get_value(row_map: dict[str, object], *keys: str) -> str:
    for key in keys:
        value = text(row_map.get(normalize(key), ""))
        if value:
            return value
    return ""


def parse_year(value: object) -> int | None:
    raw = text(value)
    match = re.search(r"\b(19|20)\d{2}\b", raw)
    if match:
        return int(match.group(0))
    try:
        return datetime.fromisoformat(raw.replace("Z", "")).year
    except ValueError:
        return None


def is_book(type_value: str) -> bool:
    key = type_value.lower()
    return any(token in key for token in ("novel", "book", "prose", "short story", "short fiction", "anthology", "collection"))


def score_candidate(query_title: str, query_year: int | None, item: dict) -> float:
    matched_title = text(item.get("title") or item.get("subtitle"))
    if not matched_title:
        return 0.0
    title_score = SequenceMatcher(None, normalize(query_title), normalize(matched_title)).ratio()
    matched_year = item.get("first_publish_year")
    year_score = 0.5
    if query_year and matched_year:
        year_score = max(0.0, 1.0 - abs(query_year - int(matched_year)) / 20.0)
    return (0.8 * title_score) + (0.2 * year_score)


def search_books(session: requests.Session, title: str, year: int | None) -> list[dict]:
    params = {"title": title, "limit": 10, "fields": "key,title,subtitle,first_publish_year,cover_i,isbn"}
    response = session.get(
        SEARCH_URL,
        params=params,
        headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
        timeout=30,
    )
    response.raise_for_status()
    return response.json().get("docs") or []


def fetch_library_url(session: requests.Session, library_url: str) -> dict:
    parsed = urlparse(library_url)
    host = parsed.netloc.lower().split(":", 1)[0]
    if host == "covers.openlibrary.org":
        cover_match = re.search(r"/b/id/(\d+)", parsed.path)
        if cover_match:
            return {"key": library_url, "cover_i": cover_match.group(1)}
    if host not in {"openlibrary.org", "www.openlibrary.org"}:
        raise ValueError("library URL must point to openlibrary.org")

    path_match = re.search(r"/(works|books|isbn)/([^/]+)", parsed.path, re.I)
    if not path_match:
        raise ValueError("library URL must contain an Open Library work, book, or ISBN path")
    resource, identifier = path_match.groups()
    response = session.get(
        f"https://openlibrary.org/{resource.lower()}/{identifier}.json",
        headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
        timeout=30,
    )
    response.raise_for_status()
    item = response.json()
    covers = item.get("covers") or []
    if covers:
        item["cover_i"] = covers[0]
    item["key"] = item.get("key") or f"/{resource.lower()}/{identifier}"
    if resource.lower() == "works":
        item["first_publish_year"] = parse_year(item.get("first_publish_date"))
    else:
        item["first_publish_year"] = parse_year(item.get("publish_date"))
    return item


def choose_result(results: list[dict], title: str, year: int | None) -> tuple[dict, float]:
    ranked = sorted(
        ((item, score_candidate(title, year, item)) for item in results),
        key=lambda pair: pair[1],
        reverse=True,
    )
    with_covers = [pair for pair in ranked if text(pair[0].get("cover_i"))]
    return (with_covers or ranked or [({}, 0.0)])[0]


def is_confident_match(score: float, query_year: int | None, matched_year: int | None, has_library_url: bool) -> bool:
    if has_library_url:
        return score >= 1.0
    if score < 0.8:
        return False
    if query_year is None or matched_year is None:
        return True
    return abs(query_year - matched_year) <= YEAR_TOLERANCE


def process_workbook(path: Path, session: requests.Session, write_posters: bool, overwrite: bool, cache: dict) -> list[dict]:
    suggestions = []
    workbook = load_workbook(path, data_only=False)
    changed = False

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=False))
        if not rows:
            continue
        headers = [text(cell.value) for cell in rows[0]]
        header_keys = [normalize(value) for value in headers]
        poster_index = next((i for i, key in enumerate(header_keys) if key == "poster"), None)
        if poster_index is None:
            continue

        for row_number, cells in enumerate(rows[1:], start=2):
            row_map = {header_keys[i]: cells[i].value for i in range(min(len(header_keys), len(cells)))}
            type_value = get_value(row_map, "type")
            if not is_book(type_value):
                continue
            existing_poster = get_value(row_map, "poster")
            if existing_poster and not overwrite:
                continue

            title = get_value(row_map, "episode title", "title", "serial title")
            release_date = get_value(row_map, "release date", "publication date", "published")
            year = parse_year(release_date)
            result = {
                "workbook": path.name, "sheet": worksheet.title, "row_number": row_number,
                "title": title, "type": type_value, "release_date": release_date,
                "matched_title": "", "matched_year": "", "openlibrary_key": "",
                "library_url": get_value(row_map, "library url", "open library url", "openlibrary url"),
                "cover_id": "", "score": 0.0, "status": "no_match", "error": "", "poster_url": "",
            }
            if not title:
                result["status"] = "missing_title"
                suggestions.append(result)
                continue

            library_url = result["library_url"]
            cache_key = ("url", library_url) if library_url else ("title", title.lower(), year)
            try:
                if library_url:
                    if cache_key not in cache:
                        try:
                            cache[cache_key] = [fetch_library_url(session, library_url)]
                        except (requests.RequestException, ValueError):
                            cache[cache_key] = []
                        time.sleep(0.1)
                    item = cache[cache_key][0] if cache[cache_key] else {}
                    score = 1.0 if item else 0.0
                    if not text(item.get("cover_i")):
                        item, score = choose_result(search_books(session, title, year), title, year)
                else:
                    if cache_key not in cache:
                        cache[cache_key] = search_books(session, title, year)
                        time.sleep(0.1)
                    item, score = choose_result(cache[cache_key], title, year)
                cover_id = text(item.get("cover_i"))
                matched_year = parse_year(item.get("first_publish_year"))
                matched = bool(cover_id) and is_confident_match(score, year, matched_year, bool(library_url))
                result.update(
                    matched_title=text(item.get("title")), matched_year=text(item.get("first_publish_year")),
                    openlibrary_key=text(item.get("key")), cover_id=cover_id, score=round(score, 4),
                    status="matched" if matched else ("review" if item else "no_match"),
                )
                if cover_id:
                    image_url = COVER_URL.format(cover_id=cover_id)
                    result["poster_url"] = image_url
                    if matched:
                        print(f"FOUND {path.name} / {worksheet.title}!{row_number} {title}: {image_url}")
                    if write_posters and matched:
                        cells[poster_index].value = image_url
                        changed = True
            except (requests.RequestException, ValueError) as exc:
                result["status"] = "error"
                result["error"] = str(exc)
            suggestions.append(result)

    if changed:
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        workbook.save(path)
    return suggestions


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", action="append", help="Workbook filename; may be repeated. Defaults to all configured workbooks.")
    parser.add_argument("--write-posters", action="store_true", help="Write Open Library cover URLs into poster cells")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing poster values")
    parser.add_argument("--csv", type=Path, default=ROOT / "openlibrary_cover_suggestions.csv")
    parser.add_argument("--json", type=Path, default=ROOT / "openlibrary_cover_suggestions.json")
    args = parser.parse_args()

    session = requests.Session()
    cache: dict = {}
    suggestions = []
    for filename in args.workbook or WORKBOOK_FILES:
        path = ROOT / filename
        if path.exists():
            suggestions.extend(process_workbook(path, session, args.write_posters, args.overwrite, cache))

    fields = list(suggestions[0].keys()) if suggestions else ["workbook", "sheet", "row_number", "status"]
    with args.csv.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        writer.writerows(suggestions)
    args.json.write_text(json.dumps(suggestions, ensure_ascii=False, indent=2), encoding="utf-8")

    counts = {status: sum(row["status"] == status for row in suggestions) for status in {row["status"] for row in suggestions}}
    print(f"Processed {len(suggestions)} published-book rows")
    print("; ".join(f"{key}: {value}" for key, value in sorted(counts.items())))
    print(f"CSV: {args.csv}")
    print(f"JSON: {args.json}")
    if not args.write_posters:
        print("Dry run only. Add --write-posters after reviewing the affected rows.")


if __name__ == "__main__":
    main()