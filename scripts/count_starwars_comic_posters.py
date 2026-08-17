from openpyxl import load_workbook

path = r"c:\Users\benja\timeline-site\STARWARS.xlsx"
wb = load_workbook(path)
ws = wb["CANON"]
headers = [c.value for c in ws[1]]
type_idx = headers.index("type")
poster_idx = headers.index("poster")

comic_rows = 0
missing = []
for row in ws.iter_rows(min_row=2, values_only=False):
    typ = row[type_idx].value if type_idx < len(row) else None
    poster = row[poster_idx].value if poster_idx < len(row) else None
    if typ and "comic" in str(typ).lower():
        comic_rows += 1
        if poster is None or str(poster).strip() == "":
            missing.append((row[3].value, typ))

print("comic_rows", comic_rows)
print("missing_among_comic", len(missing))
print("sample_missing", missing[:20])
