"""Resolve Comic Vine covers offline and optionally write local poster paths to workbooks."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import time
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CSV = ROOT / "comicvine_cover_suggestions.csv"
DEFAULT_JSON = ROOT / "comicvine_cover_suggestions.json"
WORKBOOK_FILES = [
    "MCU.xlsx",
    "STARWARS.xlsx",
    "STARTREK.xlsx",
    "DOCTORWHO.xlsx",
    "MIDDLEEARTH.xlsx",
    "RIORDANVERSE.xlsx",
    "DCU.xlsx",
    "ZELDA.xlsx",
    "ANDYWEIR.xlsx",
    "BIGBANGTHEORY.xlsx",
]
API_BASE = "https://comicvine.gamespot.com/api"
USER_AGENT = "The-Canon-Catalog Comic Vine cover resolver/1.0"


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def text(value: object) -> str:
    return str(value or "").strip()


def get_value(row_map: dict[str, object], *keys: str) -> str:
    for key in keys:
        value = text(row_map.get(normalize(key), ""))
        if value:
            return value
    return ""


def is_comic(type_value: str) -> bool:
    return "comic" in type_value.lower()


def parse_issue_number(value: str) -> str:
    match = re.search(r"\bissue\s*#?\s*0*(\d+)\b|#\s*0*(\d+)\b", value, re.I)
    if not match:
        return ""
    return next(group for group in match.groups() if group)


def parse_numeric_id(value: str) -> str:
    raw = value.strip()
    if raw.isdigit():
        return raw
    match = re.search(r"(?:-|/)(\d+)(?:/)?(?:\?.*)?$", raw)
    return match.group(1) if match else ""


def parse_directory(value: str) -> tuple[str, str]:
    raw = value.strip()
    if not raw:
        return "", ""
    numeric_id = parse_numeric_id(raw)
    if not numeric_id:
        return "", ""
    if raw.isdigit():
        return "", numeric_id
    path_parts = [part for part in urlparse(raw).path.split("/") if part]
    resource = path_parts[0] if path_parts else ""
    return resource, numeric_id


def row_title(row_map: dict[str, object]) -> str:
    return get_value(row_map, "episode title", "title", "serial title")


def row_query(row_map: dict[str, object]) -> str:
    title = row_title(row_map)
    issue = parse_issue_number(get_value(row_map, "episode"))
    return f"{title} issue {issue}" if issue else title


def comic_vine_request(session: requests.Session, api_key: str, path: str, params: dict[str, str]) -> dict:
    request_params = {"api_key": api_key, "format": "json", **params}
    response = session.get(
        f"{API_BASE}{path}",
        params=request_params,
        headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("status_code") not in (None, 1):
        raise RuntimeError(payload.get("error") or "Comic Vine request failed")
    return payload


def get_item(session: requests.Session, api_key: str, resource: str, item_id: str) -> dict:
    safe_resource = resource if resource in {"issues", "issue", "volumes", "volume"} else "issues"
    if safe_resource.endswith("s"):
        safe_resource = safe_resource[:-1]
    return comic_vine_request(session, api_key, f"/{safe_resource}/{item_id}/", {}) .get("results") or {}


def search_items(session: requests.Session, api_key: str, query: str) -> list[dict]:
    payload = comic_vine_request(
        session,
        api_key,
        "/search/",
        {"query": query, "resources": "issue,volume", "limit": "10"},
    )
    return payload.get("results") or []


def score_result(query: str, item: dict, issue_number: str) -> float:
    query_words = set(re.findall(r"[a-z0-9]+", query.lower()))
    name = text(item.get("name"))
    result_words = set(re.findall(r"[a-z0-9]+", name.lower()))
    overlap = len(query_words & result_words) / max(1, len(query_words))
    score = overlap
    if issue_number and text(item.get("issue_number")) == issue_number:
        score += 0.35
    if item.get("image", {}).get("original_url"):
        score += 0.1
    return score


def choose_result(results: list[dict], query: str, issue_number: str) -> tuple[dict, float]:
    ranked = sorted(
        ((item, score_result(query, item, issue_number)) for item in results),
        key=lambda pair: pair[1],
        reverse=True,
    )
    return ranked[0] if ranked else ({}, 0.0)


def process_workbook(path: Path, session: requests.Session, api_key: str, write_posters: bool, overwrite: bool, cache: dict) -> list[dict]:
    suggestions = []
    workbook = load_workbook(path, data_only=False)
    changed = False

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=False))
        if not rows:
            continue
        headers = [text(cell.value) for cell in rows[0]]
        header_keys = [normalize(value) for value in headers]
        poster_index = next((i for i, key in enumerate(header_keys) if key == "poster"), None)
        if poster_index is None:
            continue

        for row_number, cells in enumerate(rows[1:], start=2):
            row_map = {header_keys[i]: cells[i].value for i in range(min(len(header_keys), len(cells)))}
            type_value = get_value(row_map, "type")
            if not is_comic(type_value):
                continue

            existing_poster = get_value(row_map, "poster")
            if existing_poster and not overwrite:
                continue

            directory = get_value(row_map, "comic vine directory", "comicvine directory", "comic vine url", "comicvine url")
            comic_vine_id = get_value(row_map, "comic vine id", "comicvine id", "comic vine issue id", "comicvine issue id")
            resource, item_id = parse_directory(directory)
            item = {}
            query = row_query(row_map)
            issue_number = parse_issue_number(get_value(row_map, "episode"))
            status = "no_match"
            score = 0.0
            error = ""

            try:
                if item_id:
                    item = get_item(session, api_key, resource or "issues", item_id)
                    status = "explicit_id"
                    score = 1.0
                elif comic_vine_id:
                    item = get_item(session, api_key, "issues", parse_numeric_id(comic_vine_id))
                    status = "explicit_id"
                    score = 1.0
                elif query:
                    cache_key = (query.lower(), issue_number)
                    if cache_key not in cache:
                        cache[cache_key] = search_items(session, api_key, query)
                        time.sleep(0.35)
                    item, score = choose_result(cache[cache_key], query, issue_number)
                    status = "matched" if score >= 0.8 else ("review" if item else "no_match")
            except Exception as exc:
                error = str(exc)
                status = "error"

            image_url = text((item.get("image") or {}).get("original_url"))
            result = {
                "workbook": path.name,
                "sheet": worksheet.title,
                "row_number": row_number,
                "title": row_title(row_map),
                "type": type_value,
                "query": query,
                "comic_vine_directory": text(item.get("site_detail_url")) or directory,
                "comic_vine_id": text(item.get("id")) or item_id,
                "matched_title": text(item.get("name")),
                "issue_number": text(item.get("issue_number")),
                "image_url": image_url,
                "score": round(score, 4),
                "status": status,
                "error": error,
                "poster_url": "",
            }

            if image_url and item.get("id"):
                if write_posters:
                    cells[poster_index].value = image_url
                    result["poster_url"] = image_url
                    changed = True
                else:
                    result["poster_url"] = image_url

            suggestions.append(result)

    if changed:
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        workbook.save(path)
    return suggestions


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--api-key", default=os.getenv("COMIC_VINE_API_KEY"), help="Comic Vine API key; defaults to COMIC_VINE_API_KEY")
    parser.add_argument("--write-posters", action="store_true", help="Write Comic Vine remote image URLs into poster cells")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing poster values")
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    args = parser.parse_args()

    if not args.api_key:
        raise SystemExit("Provide --api-key or set COMIC_VINE_API_KEY; the key is never stored in the workbook or index.html")

    session = requests.Session()
    cache: dict = {}
    all_suggestions = []
    for filename in WORKBOOK_FILES:
        workbook_path = ROOT / filename
        if workbook_path.exists():
            all_suggestions.extend(process_workbook(workbook_path, session, args.api_key, args.write_posters, args.overwrite, cache))

    fields = list(all_suggestions[0].keys()) if all_suggestions else ["workbook", "sheet", "row_number", "status"]
    with args.csv.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        writer.writerows(all_suggestions)
    args.json.write_text(json.dumps(all_suggestions, ensure_ascii=False, indent=2), encoding="utf-8")

    counts = {status: sum(row["status"] == status for row in all_suggestions) for status in {row["status"] for row in all_suggestions}}
    print(f"Processed {len(all_suggestions)} comic rows")
    print("; ".join(f"{key}: {value}" for key, value in sorted(counts.items())))
    print(f"CSV: {args.csv}")
    print(f"JSON: {args.json}")


if __name__ == "__main__":
    main()
