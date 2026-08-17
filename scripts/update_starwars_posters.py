from openpyxl import load_workbook

WORKBOOK_PATH = r"c:\Users\benja\timeline-site\STARWARS.xlsx"

ISSUE_TO_COVER = {
    "https://comicvine.gamespot.com/star-wars-the-high-republic-the-blade-1-part-i/4000-961959/": "https://comicvine.gamespot.com/a/uploads/scale_large/11/110017/8766097-wwww.jpg",
    "https://comicvine.gamespot.com/star-wars-the-high-republic-the-blade-2-part-ii/4000-966428/": "https://comicvine.gamespot.com/a/uploads/scale_large/11/110017/8795910-wwww.jpg",
    "https://comicvine.gamespot.com/star-wars-the-high-republic-the-blade-3-part-iii/4000-974100/": "https://comicvine.gamespot.com/a/uploads/scale_large/11/110017/8841816-wwww.jpg",
    "https://comicvine.gamespot.com/star-wars-the-high-republic-the-blade-4-part-iv/4000-979435/": "https://comicvine.gamespot.com/a/uploads/scale_large/11144/111442876/8881647-preview-for-the-high-republic-the-blade-of-4-v0-0angj8puiqpa1.jpg",
}

wb = load_workbook(WORKBOOK_PATH)
ws = wb["CANON"]
headers = [cell.value for cell in ws[1]]
poster_idx = headers.index("poster")
cv_idx = headers.index("comic vine id")
updated = []

for row in ws.iter_rows(min_row=2, values_only=False):
    raw = row[cv_idx].value if cv_idx < len(row) else None
    if raw is None or not str(raw).strip():
        continue
    value = str(raw).strip()
    cover_url = ISSUE_TO_COVER.get(value)
    if not cover_url:
        continue
    row[poster_idx].value = cover_url
    updated.append((row[3].value, cover_url))
    print("UPDATED", row[3].value, "->", cover_url)

print("TOTAL_UPDATED", len(updated))
wb.save(WORKBOOK_PATH)
