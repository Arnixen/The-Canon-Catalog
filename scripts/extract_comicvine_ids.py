"""Extract Comic Vine issue IDs from workbook URL columns without network access."""

from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "STARWARS.xlsx"


def text(value: object) -> str:
    return str(value or "").strip()


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def extract_issue_id(value: object) -> str:
    raw = text(value)
    if not raw:
        return ""
    path = urlparse(raw).path if re.match(r"https?://", raw, re.I) else raw
    match = re.search(r"(?:^|/)(?:[^/]*-)?(\d+)(?:/)?$", path)
    return match.group(1) if match else ""


def process_workbook(path: Path, write: bool, overwrite: bool) -> tuple[int, int, int]:
    workbook = load_workbook(path, data_only=False)
    scanned = written = skipped = 0
    changed = False

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=False))
        if not rows:
            continue
        headers = [normalize(cell.value) for cell in rows[0]]
        url_index = next((i for i, key in enumerate(headers) if key == "comicvineurl"), None)
        id_index = next((i for i, key in enumerate(headers) if key == "comicvineid"), None)
        if url_index is None or id_index is None:
            continue

        for row_number, cells in enumerate(rows[1:], start=2):
            issue_url = text(cells[url_index].value)
            if not issue_url:
                continue
            scanned += 1
            issue_id = extract_issue_id(issue_url)
            if not issue_id:
                print(f"NO_ID {worksheet.title}!{row_number}: {issue_url}")
                continue
            if text(cells[id_index].value) and not overwrite:
                skipped += 1
                continue
            print(f"FOUND {worksheet.title}!{row_number}: {issue_id}")
            if write:
                cells[id_index].value = issue_id
                written += 1
                changed = True

    if changed:
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        workbook.save(path)
    return scanned, written, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--write", action="store_true", help="Write extracted IDs into the workbook")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing Comic Vine ID values")
    args = parser.parse_args()

    scanned, written, skipped = process_workbook(args.workbook, args.write, args.overwrite)
    print(f"Rows with Comic Vine URLs: {scanned}")
    print(f"IDs {'written' if args.write else 'found'}: {written if args.write else scanned - skipped}")
    print(f"Existing IDs skipped: {skipped}")
    if not args.write:
        print("Dry run only. Add --write to update the workbook.")


if __name__ == "__main__":
    main()