"""Populate poster cells from public Comic Vine issue pages without API requests."""

from __future__ import annotations

import argparse
import html
import re
import shutil
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "STARWARS.xlsx"
USER_AGENT = "The-Canon-Catalog Comic Vine page resolver/1.0"


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def text(value: object) -> str:
    return str(value or "").strip()


class OpenGraphParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.image_url = ""

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "meta" or self.image_url:
            return
        attributes = {name.lower(): text(value) for name, value in attrs}
        property_name = attributes.get("property", "").lower()
        if property_name == "og:image":
            self.image_url = html.unescape(attributes.get("content", ""))


def cover_url_from_page(session: requests.Session, issue_url: str) -> str:
    response = session.get(
        issue_url,
        headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml"},
        timeout=30,
    )
    response.raise_for_status()
    parser = OpenGraphParser()
    parser.feed(response.text)
    return urljoin(response.url, parser.image_url)


def process_workbook(path: Path, sheet_name: str, write_posters: bool, overwrite: bool) -> tuple[int, int, int]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=False))
    if not rows:
        return 0, 0, 0

    headers = [text(cell.value) for cell in rows[0]]
    keys = [normalize(value) for value in headers]
    poster_index = keys.index("poster")
    comic_vine_index = keys.index("comicvineid")
    processed = written = failed = 0
    changed = False
    session = requests.Session()

    for row_number, cells in enumerate(rows[1:], start=2):
        issue_url = text(cells[comic_vine_index].value)
        existing_poster = text(cells[poster_index].value)
        if not issue_url or (existing_poster and not overwrite):
            continue

        processed += 1
        try:
            cover_url = cover_url_from_page(session, issue_url)
            if not cover_url:
                raise ValueError("page did not contain an og:image cover URL")
            print(f"FOUND row {row_number}: {cover_url}")
            if write_posters:
                cells[poster_index].value = cover_url
                changed = True
                written += 1
        except (requests.RequestException, ValueError) as exc:
            failed += 1
            print(f"FAILED row {row_number}: {exc}")

    if changed:
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        workbook.save(path)
    return processed, written, failed


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default="CANON")
    parser.add_argument("--write-posters", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    processed, written, failed = process_workbook(
        args.workbook,
        args.sheet,
        args.write_posters,
        args.overwrite,
    )
    print(f"Rows with Comic Vine URLs processed: {processed}")
    print(f"Poster URLs {'written' if args.write_posters else 'found'}: {written if args.write_posters else processed - failed}")
    print(f"Lookup failures: {failed}")
    if not args.write_posters:
        print("Dry run only. Add --write-posters after reviewing the results.")


if __name__ == "__main__":
    main()